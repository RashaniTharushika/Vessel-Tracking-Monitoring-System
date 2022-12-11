import json
import pandas as pd
import requests
from datetime import datetime
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

now = datetime.now()

Endpoint_Name = 'SINGLE_VESSEL_POSITION'
print(now)
SINGLE_VESSEL_POSITIONING_API_KEY = 'fa707edb08ff303c0d77f6a8cc40b239401c5a49'


# def fill_nan(filepath, sheet_name='Sheet1'):
#     dataframe = pd.read_excel(filepath, sheet_name=sheet_name)

#     for column in dataframe.columns:
#         dataframe[column].fillna("Not Found", inplace=True)

#     book = load_workbook(filepath)
#     writer = pd.ExcelWriter(filepath, engine='openpyxl')
#     writer.book = book

#     writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

#     dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

#     writer.close()

def single_vessel_position(FILE_PATH):
    FILEPATH = f"{FILE_PATH}/Master file - Input file/Master File.xlsx"
    SHEET_NAME = 'Master File_To be Daily updated'

    df = pd.read_excel(FILEPATH, sheet_name=SHEET_NAME)

    df = df.loc[df['MBL NO'] == 'MBL NOT GENERATED']

    FILE_PATH_errorlog = f"{FILE_PATH}/API Output Files/Error Log.xlsx"
    SHEET_NAME_errorlog = 'Sheet1'

    df_error_log = pd.read_excel(FILE_PATH_errorlog, sheet_name=SHEET_NAME_errorlog)

    status_dict = {
        '0': 'Under way using engine',
        '1': 'At anchor',
        '2': 'Not under command',
        '3': 'Restricted manoeuverability',
        '4': 'Constrained by her draught',
        '5': 'Moored',
        '6': 'Aground',
        '7': 'Engaged in Fishing',
        '8': 'Under way sailing',
        '9': 'Reserved for future amendment of Navigational Status for HSC',
        '10': 'Reserved for future amendment of Navigational Status for WIG',
        '11': 'Reserved for future use',
        '12': 'Reserved for future use',
        '13': 'Reserved for future use',
        '14': 'AIS-SART is active',
        '15': 'Not defined (default)'

    }

    print(df)
    df = df.reindex(
        columns=[*df.columns.tolist(),
                 'latitude',
                 'longitude',
                 'speed',
                 'heading',
                 'course',
                 'status',
                 'Time Stamp',
                 'Shipname',
                 'Ship Type',
                 'Type Name',
                 'AIS type',
                 'IMO',
                 'Callsign',
                 'Flag',
                 'port ID',
                 'Port unlocode',
                 'Current port',
                 'last port ID',
                 'Last port unlocode',
                 'last port',
                 'Last port time',
                 'destination',
                 'ETA',
                 'ETA Calc',
                 'Length',
                 'Width',
                 'Draught',
                 'GRT',
                 'Next port ID',
                 'Next port unlocode',
                 'Next port Name',
                 'Next port country',
                 'DWT',
                 'Year Build',
                 'DSRC',
                 'last_modified_time'
                 ],
        fill_value=None)

    Timestamp = 300
    vessel_pos = 0

    for i in range(0, len(df)):
        try:
            vessel_pos = vessel_pos + 1
            url = "https://services.marinetraffic.com/api/exportvessel/" + SINGLE_VESSEL_POSITIONING_API_KEY + "/v:5/mmsi:" + str(
                int(df.iloc[i][11])) + "/timespan:" + str(Timestamp) + "/protocol:jsono/msgtype:extended"

            print(url)
            ## ERROR - requests.exceptions.ConnectionError: ('Connection aborted.', OSError(0, 'Error'))
            ## If the provided API credit are over would raise above errror.

            response = requests.get(url)
            print(response)
            json_response = response.json()
            print(json_response)
            if response.status_code == 200:
                response_SVP = json_response[0]
                print(response_SVP)

                latitude = response_SVP['LAT']
                longitude = response_SVP['LON']
                speed = response_SVP['SPEED']
                heading = response_SVP['HEADING']
                course = response_SVP['COURSE']
                status = response_SVP['STATUS']
                last_reported_time = response_SVP['TIMESTAMP']
                ship_name = response_SVP['SHIPNAME']
                ship_type = response_SVP['SHIPTYPE']
                type_name = response_SVP['TYPE_NAME']
                AIS_type = response_SVP['AIS_TYPE_SUMMARY']
                Imo = response_SVP['IMO']
                Callsign = response_SVP['CALLSIGN']
                flag = response_SVP['FLAG']
                port_ID = response_SVP["PORT_ID"]
                Port_unlocode = response_SVP['PORT_UNLOCODE']
                current_port = response_SVP['CURRENT_PORT']
                last_portID = response_SVP['LAST_PORT_ID']
                Last_port_unlocode = response_SVP['LAST_PORT_UNLOCODE']
                last_port = response_SVP["LAST_PORT"]
                Last_port_time = response_SVP['LAST_PORT_TIME']
                destination = response_SVP['DESTINATION']
                ETA_AIS = response_SVP['ETA']
                ETA_Calc = response_SVP["ETA_CALC"]
                Length = response_SVP['LENGTH']
                width = response_SVP['WIDTH']
                draught = response_SVP['DRAUGHT']
                grt = response_SVP['GRT']
                Next_port_ID = response_SVP['NEXT_PORT_ID']
                Next_port_unlocode = response_SVP['NEXT_PORT_UNLOCODE']
                Next_port_Name = response_SVP['NEXT_PORT_NAME']
                Next_port_country = response_SVP['NEXT_PORT_COUNTRY']
                Dwt = response_SVP['DWT']
                year = response_SVP['YEAR_BUILT']
                dsrc = response_SVP['DSRC']

                print(f'******** {status}')
                status_up = [status_dict.get(key) for key in status]
                # print(lastPort,lastPort_time,nextPort,calculated_ETA_to_nextPort)

                df.iloc[i, 15] = latitude
                df.iloc[i, 16] = longitude
                df.iloc[i, 17] = speed
                df.iloc[i, 18] = heading
                df.iloc[i, 19] = course
                df.iloc[i, 20] = status_up
                df.iloc[i, 21] = last_reported_time
                df.iloc[i, 22] = ship_name
                df.iloc[i, 23] = ship_type
                df.iloc[i, 24] = type_name
                df.iloc[i, 25] = AIS_type
                df.iloc[i, 26] = Imo
                df.iloc[i, 27] = Callsign
                df.iloc[i, 28] = flag
                df.iloc[i, 29] = port_ID
                df.iloc[i, 30] = Port_unlocode
                df.iloc[i, 31] = current_port
                df.iloc[i, 32] = last_portID
                df.iloc[i, 33] = Last_port_unlocode
                df.iloc[i, 34] = last_port
                df.iloc[i, 35] = Last_port_time
                df.iloc[i, 36] = destination
                df.iloc[i, 37] = ETA_AIS
                df.iloc[i, 38] = ETA_Calc
                df.iloc[i, 39] = Length
                df.iloc[i, 40] = width
                df.iloc[i, 41] = draught
                df.iloc[i, 42] = grt
                df.iloc[i, 43] = Next_port_ID
                df.iloc[i, 44] = Next_port_unlocode
                df.iloc[i, 45] = Next_port_Name
                df.iloc[i, 46] = Next_port_country
                df.iloc[i, 47] = Dwt
                df.iloc[i, 48] = year
                df.iloc[i, 49] = dsrc
                df.iloc[i, 50] = datetime.now()

            else:
                error_response_dict = {
                    'Status': 'error',
                    'Code': json_response['errors'][0]['code'],
                    'Message': json_response['errors'][0]['detail'],
                    'Time': now,
                    'MBL': "MMSI No is " + str(int(df.iloc[i][11])),
                    'ShipmentID': "Generated when tracking vessels before BL cutoff, no shipment ID"
                }

                df_error_log = df_error_log.append(error_response_dict, ignore_index=True)

        except Exception as e:
            pass

    df_error_log.to_excel(f"{FILE_PATH}/API Output Files/Error Log.xlsx", index=False)

    df.to_excel(f'{FILE_PATH}/API Output Files/SVP.xlsx', index=False)

    # fill_nan('C:/Users/User/MAS Holdings (Pvt) Ltd/Logistics DA Project - Documents/General/API Output Files/SVP.xlsx')

    # with pd.ExcelWriter("C:/Users/User/MAS Holdings (Pvt) Ltd/Logistics DA Project - Documents/General/Test/ETA_TO_PORT/SINGLE_VESSEL_POSITIONING.xlsx",mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
    #     df.to_excel(writer, sheet_name="Sheet1",header=None, startrow=writer.sheets["Sheet1"].max_row,index=False)

    # df1=pd.read_excel("SINGLE_VESSEL_POSITIONING.xlsx")

    # frames = [df1,df]

    # Full_df = pd.concat(frames)
    # print(Full_df)
    # Full_df.to_excel('C:/Users/User/MAS Holdings (Pvt) Ltd/Logistics DA Project - Documents/General/Test/SINGLE_VESSEL_POSITION/SINGLE_VESSEL_POSITIONING.xlsx', index=False)

    print('COMPLETE')
    return vessel_pos


# if __name__ == '__main__':
#     single_vessel_position()
